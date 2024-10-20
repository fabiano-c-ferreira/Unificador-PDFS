import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
from PyPDF2 import PdfMerger
from datetime import datetime
from tkinter import ttk

CONFIG_FILE = 'config.json'

# Mapeamento de meses em inglês para português
meses_map = {
    "January": "Janeiro", "February": "Fevereiro", "March": "Março",
    "April": "Abril", "May": "Maio", "June": "Junho",
    "July": "Julho", "August": "Agosto", "September": "Setembro",
    "October": "Outubro", "November": "Novembro", "December": "Dezembro"
}

# Função para carregar as configurações salvas
def carregar_configuracoes():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r') as file:
            return json.load(file)
    return {'mes': 'Janeiro', 'ano': str(datetime.now().year)}  # Valor padrão

# Função para salvar as configurações atuais
def salvar_configuracoes(pasta, planilha, mes, ano):
    with open(CONFIG_FILE, 'w') as file:
        json.dump({'pasta': pasta, 'planilha': planilha, 'mes': mes, 'ano': ano}, file)

# Função para selecionar a pasta de PDFs
def selecionar_pasta():
    pasta = filedialog.askdirectory()
    if pasta:
        entrada_pasta.set(pasta)
        salvar_configuracoes(pasta, entrada_planilha.get(), mes_selecionado.get(), ano_selecionado.get())

# Função para selecionar o arquivo Excel
def selecionar_planilha():
    arquivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if arquivo:
        entrada_planilha.set(arquivo)
        salvar_configuracoes(entrada_pasta.get(), arquivo, mes_selecionado.get(), ano_selecionado.get())

# Função para obter a lista de nomes a partir da aba selecionada
def obter_lista_nomes(caminho_planilha, aba_selecionada, num_arquivos):
    workbook = openpyxl.load_workbook(caminho_planilha)
    sheet = workbook[aba_selecionada]
    # Pegar apenas as últimas 'num_arquivos' linhas
    lista_nomes = [sheet['A'][i].value.strip() for i in range(sheet.max_row - num_arquivos, sheet.max_row) if sheet['A'][i].value is not None]
    workbook.close()
    return lista_nomes

# Função para renomear os arquivos na pasta
def renomear_arquivos(pasta, lista_nomes, ano, mes):
    lista_arquivos = [os.path.join(pasta, arquivo) for arquivo in os.listdir(pasta) if arquivo.endswith('.pdf')]
    lista_arquivos.sort(key=os.path.getctime)
    lista_arquivos.reverse()  # inverte a lista de nomes para nomear os pdfs

    total = len(lista_arquivos)
    for i, caminho_antigo in enumerate(lista_arquivos):
        if i < len(lista_nomes):
            nome_cliente = lista_nomes[i]
            novo_nome = f'NFS-e {ano}.{mes:02d} - {nome_cliente}.pdf'
            caminho_novo = os.path.join(pasta, novo_nome)
            try:
                if os.path.isfile(caminho_antigo):
                    os.rename(caminho_antigo, caminho_novo)
                else:
                    print(f"Arquivo não encontrado: {caminho_antigo}")
            except Exception as e:
                print(f"Erro ao renomear {os.path.basename(caminho_antigo)}: {e}")
        
        # Atualizar a barra de progresso
        progress_bar['value'] = (i + 1) / total * 50  # 50% para renomeação
        janela.update_idletasks()

def mesclar_pdfs(pasta_entrada, arquivo_saida, lista_nomes):
    merger = PdfMerger()
    nome_para_caminho = {}
    lista_nomes.reverse()  # inverte a lista de nomes para acertar a ordem dos pdfs mesclados
    for nome in lista_nomes:
        nome_arquivo = f'NFS-e {ano_selecionado.get()}.{list(meses_map.values()).index(mes_selecionado.get()) + 1:02d} - {nome}.pdf'
        caminho_arquivo = os.path.join(pasta_entrada, nome_arquivo)
        if os.path.exists(caminho_arquivo):
            nome_para_caminho[nome] = caminho_arquivo

    total = len(lista_nomes)
    for i, nome in enumerate(lista_nomes):
        caminho_arquivo = nome_para_caminho.get(nome)
        if caminho_arquivo:
            merger.append(caminho_arquivo)

        # Atualizar a barra de progresso
        progress_bar['value'] = 50 + (i + 1) / total * 45  # 45% para mesclagem
        janela.update_idletasks()

    merger.write(arquivo_saida)
    merger.close()
    
    # Finalizar a barra de progresso
    progress_bar['value'] = 100
    janela.update_idletasks()

# Função para executar o processo
def executar_processo():
    pasta = entrada_pasta.get()
    planilha = entrada_planilha.get()
    mes = mes_selecionado.get()
    ano = ano_selecionado.get()

    if not pasta or not planilha:
        messagebox.showerror("Erro", "Por favor, selecione a pasta e a planilha.")
        return

    # Verificar se o mês ou ano selecionado é diferente do atual
    mes_atual_ingles = datetime.now().strftime("%B")
    mes_atual_portugues = meses_map[mes_atual_ingles]
    ano_atual = str(datetime.now().year)
    if mes != mes_atual_portugues or ano != ano_atual:
        resposta = messagebox.askyesno("Confirmação", f"O mês ou ano selecionado ({mes}/{ano}) é diferente do atual ({mes_atual_portugues}/{ano_atual}). Deseja continuar?")
        if not resposta:
            return

    # Converter o mês para número
    mes_numero = list(meses_map.values()).index(mes) + 1

    lista_arquivos = [arquivo for arquivo in os.listdir(pasta) if arquivo.endswith('.pdf')]
    num_arquivos = len(lista_arquivos)
    lista_nomes = obter_lista_nomes(planilha, mes, num_arquivos)
    renomear_arquivos(pasta, lista_nomes, ano, mes_numero)
    arquivo_saida = os.path.join(pasta, f'PDFao_Mescladao.pdf')
    mesclar_pdfs(pasta, arquivo_saida, lista_nomes)

    salvar_configuracoes(pasta, planilha, mes, ano)
    messagebox.showinfo("Concluído", "Processo concluído com sucesso!")

# Configuração da janela principal
janela = tk.Tk()
janela.title("Renomeador de NFS-e - V 1.00")
janela.iconbitmap("icone-heat.ico")
janela.geometry("400x480")

# Carregar configurações ao iniciar
configuracoes = carregar_configuracoes()
entrada_pasta = tk.StringVar(value=configuracoes.get('pasta', ''))
entrada_planilha = tk.StringVar(value=configuracoes.get('planilha', ''))
mes_selecionado = tk.StringVar(value=configuracoes.get('mes', 'Janeiro'))
ano_selecionado = tk.StringVar(value=configuracoes.get('ano', str(datetime.now().year)))

# Carregar e exibir o logotipo
try:
    imagem_logo = Image.open("logo.png")
    largura_original, altura_original = imagem_logo.size
    nova_largura = int(largura_original * 0.20)
    nova_altura = int(altura_original * 0.20)
    imagem_logo = imagem_logo.resize((nova_largura, nova_altura), Image.LANCZOS)
    logo = ImageTk.PhotoImage(imagem_logo)
    label_logo = tk.Label(janela, image=logo)
    label_logo.place(relx=0.015, rely=0.015, anchor='nw')
except Exception as e:
    print(f"Erro ao carregar o logotipo: {e}")

# Adicionar título
titulo = tk.Label(janela, text="Renomeador de NFS-e", font=("Segoe UI", 17, "bold"), fg="#555555")
titulo.place(relx=0.5, rely=0.03, anchor='n')

# Dropdown para selecionar o mês e entrada para o ano
frame_meses_ano = tk.Frame(janela)
frame_meses_ano.pack(pady=(50, 20))

tk.Label(frame_meses_ano, text="Mês:", font=("Segoe UI", 11, "bold"), fg="#555555").grid(row=0, column=0, padx=(20, 2), pady=(30, 1))
dropdown_meses = tk.OptionMenu(frame_meses_ano, mes_selecionado, *['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
                                                                    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'])
dropdown_meses.grid(row=0, column=1, padx=(5, 40), pady=(30, 1))

tk.Label(frame_meses_ano, text="Ano:", font=("Segoe UI", 11, "bold"), fg="#555555").grid(row=0, column=2, padx=(20, 2), pady=(30, 1))
entry_ano = tk.Entry(frame_meses_ano, textvariable=ano_selecionado, width=5, font=("Segoe UI", 12))
entry_ano.grid(row=0, column=3, padx=(5, 20), pady=(30, 1))

# Entrada para a pasta
tk.Label(janela, text="Selecione a pasta de PDFs:", font=("Segoe UI", 11, "bold"), fg="#555555").pack(padx=(70, 5), pady=(30, 5))
frame_pasta = tk.Frame(janela)
frame_pasta.pack(padx=(10, 15))
tk.Button(frame_pasta, text="Procurar", command=selecionar_pasta).pack(side=tk.LEFT, padx=10, pady=(0,15))
tk.Entry(frame_pasta, textvariable=entrada_pasta, width=38).pack(side=tk.LEFT, pady=(0,15))

# Entrada para o arquivo Excel
tk.Label(janela, text="Selecione a planilha Excel:", font=("Segoe UI", 11, "bold"), fg="#555555").pack(padx=(70, 5), pady=(30, 5))
frame_planilha = tk.Frame(janela)
frame_planilha.pack(padx=(10, 15))
tk.Button(frame_planilha, text="Procurar", command=selecionar_planilha).pack(side=tk.LEFT, padx=10, pady=(0,15))
tk.Entry(frame_planilha, textvariable=entrada_planilha, width=38).pack(side=tk.LEFT, pady=(0,15))

# Barra de progresso
progress_bar = ttk.Progressbar(janela, orient='horizontal', length=300, mode='determinate')
progress_bar.pack(pady=(40, 10))

# Botão para executar o processo
tk.Button(janela, text="Renomear", font=("Segoe UI", 13, "bold"), command=executar_processo, bg="#ec6608", fg="white").pack(pady=(5,15))

janela.mainloop()